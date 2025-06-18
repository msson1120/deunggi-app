import streamlit as st
import pandas as pd
import tempfile
import zipfile
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="(주)건화 등기부등본 Excel 통합기", layout="wide")

password = st.text_input('비밀번호를 입력하세요', type='password')
if password != '1220':
    st.warning('올바른 비밀번호를 입력하세요.')
    st.stop()

st.title("📦 (주)건화 등기부등본 통합분석기")
st.markdown("""
압축파일(.zip) 안의 폴더 구조와 관계없이 모든 엑셀 파일을 자동 분석합니다.
""")

uploaded_zip = st.file_uploader("📁 .zip 파일을 업로드하세요 (내부에 .xlsx 파일 포함)", type=["zip"])
run_button = st.button("분석 시작")

def merge_adjacent_cells(row_series, max_gap=3):
    """
    인접한 셀들을 병합하여 하나의 의미있는 단위로 만드는 함수
    데이터 행에서는 더 신중하게 병합
    """
    merged_row = row_series.copy()
    row_dict = row_series.to_dict()
    
    # 빈 셀이 아닌 셀들의 인덱스를 찾기
    non_empty_indices = [idx for idx, val in row_dict.items() if str(val).strip()]
    
    # 데이터가 너무 적거나 많으면 병합하지 않음 (헤더가 아닌 경우)
    if len(non_empty_indices) < 2 or len(non_empty_indices) > 10:
        return merged_row
    
    # 연속된 셀들을 그룹화 (더 엄격한 조건)
    groups = []
    current_group = []
    
    for i, idx in enumerate(non_empty_indices):
        if not current_group:
            current_group = [idx]
        else:
            # 이전 인덱스와의 거리가 2 이하면 같은 그룹 (더 엄격하게)
            if idx - current_group[-1] <= 2:
                current_group.append(idx)
            else:
                # 새로운 그룹 시작
                groups.append(current_group)
                current_group = [idx]
    
    if current_group:
        groups.append(current_group)
    
    # 각 그룹 내의 셀들을 병합 (더 신중하게)
    for group in groups:
        if len(group) > 1 and len(group) <= 3:  # 너무 많은 셀은 병합하지 않음
            # 그룹 내 모든 값을 연결
            merged_value = ""
            for idx in group:
                val = str(row_dict.get(idx, "")).strip()
                if val:
                    if merged_value and not merged_value.endswith((" ", "-", "/")):
                        merged_value += " "
                    merged_value += val
            
            # 첫 번째 인덱스에 병합된 값 저장
            merged_row[group[0]] = merged_value
            
            # 나머지 인덱스는 빈 값으로 설정
            for idx in group[1:]:
                merged_row[idx] = ""
    
    return merged_row

def merge_dataframe_cells(df, is_header_row=False):
    """
    데이터프레임에 셀 병합 로직 적용
    헤더 행과 데이터 행을 구분하여 처리
    """
    if df.empty:
        return df
    
    merged_df = df.copy()
    
    # 첫 번째 행은 헤더로 가정하고 더 관대하게 병합
    if len(merged_df) > 0:
        merged_df.iloc[0] = merge_adjacent_cells(merged_df.iloc[0], max_gap=3)
    
    # 나머지 행들은 데이터 행으로 더 엄격하게 병합
    for i in range(1, len(merged_df)):
        merged_df.iloc[i] = merge_adjacent_cells(merged_df.iloc[i], max_gap=2)
    
    return merged_df

def trim_after_reference_note(df):
    for i, row in df.iterrows():
        row_text = "".join(str(cell) for cell in row)
        normalized = re.sub(r"\s+", "", row_text)
        if "참고사항" in normalized or "참고" in normalized or "비고" in normalized:
            return df.iloc[:i]
    return df

def extract_identifier(df):
    """
    파일에서 토지/건물 식별자를 추출하는 함수
    """
    for i in range(len(df)):
        row = df.iloc[i]
        row_text = " ".join(str(cell) for cell in row if pd.notna(cell))
        if "고유번호" in row_text:
            for j in range(i+1, min(i+10, len(df))):
                content = " ".join(str(cell) for cell in df.iloc[j] if pd.notna(cell))
                if content.strip().startswith(("[토지]", "[건물]")):
                    return content.strip()
            break
    
    # 고유번호 이후에 [토지] 또는 [건물]이 없는 경우, 전체 데이터에서 찾기
    for i in range(len(df)):
        row_text = " ".join(str(cell) for cell in df.iloc[i] if pd.notna(cell))
        if row_text.strip().startswith(("[토지]", "[건물]")):
            return row_text.strip()
            
    return "알수없음"

def convert_jibun_to_decimal(jibun_text):
    """
    최종지분 텍스트를 소수점 형태로 변환하는 함수
    예: "2분의 1" -> 0.5, "1/2" -> 0.5, "50%" -> 0.5, "단독소유" -> 1
    """
    if not jibun_text or pd.isna(jibun_text):
        return None
    
    jibun_text = str(jibun_text).strip()
    
    # 단독소유는 1로 변환
    if "단독소유" in jibun_text or (("단독" in jibun_text) and len(jibun_text) < 10):
        return 1.0
    
    # 1) 분수 형태 (예: 1/2, 1/3, 공유1/3 등)
    fraction_match = re.search(r'(?:공유)?(\d+)/(\d+)', jibun_text)
    if fraction_match:
        numerator = float(fraction_match.group(1))
        denominator = float(fraction_match.group(2))
        if denominator != 0:
            return numerator / denominator
    
    # 2) 퍼센트 형태 (예: 50%, 33.3% 등)
    percent_match = re.search(r'([\d\.]+)\s*%', jibun_text)
    if percent_match:
        return float(percent_match.group(1)) / 100
    
    # 3) '분의' 형태 (예: 3분의 1, 2분의 1 등)
    boonui_match = re.search(r'(\d+\.?\d*)\s*분\s*의\s*(\d+\.?\d*)', jibun_text)
    if boonui_match:
        denominator = float(boonui_match.group(1))
        numerator = float(boonui_match.group(2))
        if denominator != 0:
            return numerator / denominator
    
    # 4) 분의 형태 - 띄어쓰기 없는 경우 (예: 10139.94분의845.0298)
    boonui_match2 = re.search(r'(\d+\.?\d*)분의(\d+\.?\d*)', jibun_text)
    if boonui_match2:
        denominator = float(boonui_match2.group(1))
        numerator = float(boonui_match2.group(2))
        if denominator != 0:
            return numerator / denominator
    
    return None

def keyword_match_partial(cell, keyword):
    if pd.isnull(cell): return False
    return keyword.replace(" ", "") in str(cell).replace(" ", "")

def keyword_match_exact(cell, keyword):
    if pd.isnull(cell): return False
    return re.sub(r"\s+", "", str(cell)) == re.sub(r"\s+", "", keyword)

def merge_split_headers(header_row):
    """분리된 헤더를 병합하는 함수 - 개선된 버전"""
    # 셀 병합을 하지 않고 원본 헤더를 그대로 사용
    merged_row = header_row.copy()
    
    # 기존 특정 키워드 병합 로직만 적용 (인접 셀 병합은 제외)
    split_patterns = {
        "주소": ["주", "소"],
        "등기명의인": ["등기", "명의인"],
        "주민등록번호": ["주민", "등록번호"],
        "최종지분": ["최종", "지분"],
        "순위번호": ["순위", "번호"],
        "등기목적": ["등기", "목적"],
        "접수정보": ["접수", "정보"],
        "주요등기사항": ["주요", "등기사항"],
        "대상소유자": ["대상", "소유자"]
    }
    
    for target_keyword, split_parts in split_patterns.items():
        found_indices = []
        for part in split_parts:
            for idx, cell_value in merged_row.items():
                cell_str = str(cell_value).strip()
                if cell_str == part:
                    found_indices.append(idx)
                    break
        
        if len(found_indices) == len(split_parts):
            if all(found_indices[i+1] - found_indices[i] <= 2 for i in range(len(found_indices)-1)):
                merged_row[found_indices[0]] = target_keyword
                for idx in found_indices[1:]:
                    merged_row[idx] = ""
    
    return merged_row

def enhanced_keyword_match(header_row, keyword, max_distance=2):
    """인접한 셀들을 고려한 키워드 매칭 - 개선된 버전"""
    # 먼저 정확한 매칭 시도
    for idx, cell in header_row.items():
        if keyword_match_exact(cell, keyword):
            return idx
    
    # 부분 매칭 시도
    for idx, cell in header_row.items():
        if keyword_match_partial(cell, keyword):
            return idx
    
    # 분리된 키워드 매칭 시도 (더 엄격하게)
    keyword_chars = list(keyword.replace(" ", ""))
    if len(keyword_chars) <= 1:
        return None
    
    for start_idx, cell in header_row.items():
        if str(cell).strip() == keyword_chars[0]:
            # 첫 글자가 매칭되면 다음 글자들을 인접 셀에서 찾기
            current_text = str(cell).strip()
            current_idx = start_idx
            
            for i in range(1, len(keyword_chars)):
                found_next = False
                # 최대 max_distance까지 떨어진 셀에서 다음 글자 찾기
                for offset in range(1, max_distance + 1):
                    next_idx = current_idx + offset
                    if next_idx in header_row:
                        next_cell = str(header_row[next_idx]).strip()
                        if next_cell == keyword_chars[i]:
                            current_text += next_cell
                            current_idx = next_idx
                            found_next = True
                            break
                
                if not found_next:
                    break
            
            # 전체 키워드가 매칭되었는지 확인
            if current_text == keyword.replace(" ", ""):
                return start_idx
    
    return None

def extract_section_range(df, start_kw, end_kw_list, match_fn):
    df = df.fillna("")
    df.columns = range(df.shape[1])
    start_idx, end_idx = None, len(df)
    for i, row in df.iterrows():
        if any(match_fn(cell, start_kw) for cell in row):
            start_idx = i + 1
            break
    if start_idx is None:
        return pd.DataFrame(), False
    for i in range(start_idx, len(df)):
        row = df.iloc[i]
        if any(any(match_fn(cell, end_kw) for cell in row) for end_kw in end_kw_list):
            end_idx = i
            break
    section = df.iloc[start_idx:end_idx].copy()
    is_empty = section.replace("", pd.NA).dropna(how="all").empty
    return section if not is_empty else pd.DataFrame([["기록없음"]]), not is_empty

# 소유지분현황(갑구)에서 필요한 열을 추출
def extract_named_cols(section, col_keywords):
    if section.empty:
        return pd.DataFrame([["기록없음"]])
    
    # 셀 병합 적용 (헤더와 데이터 구분)
    section = merge_dataframe_cells(section)
    
    header_row = section.iloc[0]
    merged_header = merge_split_headers(header_row)
    
    col_map = {}
    for target in col_keywords:
        col_idx = enhanced_keyword_match(merged_header, target)
        if col_idx is not None:
            col_map[target] = col_idx

    # 최종지분 특별 처리 (기존 로직 유지하되 더 정확하게)
    if "최종지분" not in col_map:
        idx_최종 = None
        idx_지분 = None
        for idx, val in merged_header.items():
            val_str = str(val).strip()
            if val_str == "최종":
                idx_최종 = idx
            elif val_str == "지분":
                idx_지분 = idx
        
        if idx_최종 is not None and idx_지분 is not None and abs(idx_최종 - idx_지분) <= 2:
            col_map["최종지분"] = (min(idx_최종, idx_지분), max(idx_최종, idx_지분))

    rows = []
    for i in range(1, len(section)):
        row = section.iloc[i]
        row_dict = {}
        
        for key in col_keywords:
            if key == "최종지분":
                if isinstance(col_map.get("최종지분"), tuple):
                    idx1, idx2 = col_map["최종지분"]
                    val1 = str(row.get(idx1, "")).strip()
                    val2 = str(row.get(idx2, "")).strip()
                    if val1 and val2:
                        row_dict[key] = val1 + val2
                    else:
                        row_dict[key] = val1 or val2
                elif isinstance(col_map.get("최종지분"), int):
                    idx = col_map["최종지분"]
                    val1 = str(row.get(idx, "")).strip()
                    # 인접 셀 확인은 헤더가 비어있을 때만
                    val2 = ""
                    if (idx + 1) in row and not str(merged_header.get(idx + 1, "")).strip():
                        val2 = str(row.get(idx + 1, "")).strip()
                    if val1 and val2:
                        row_dict[key] = val1 + val2
                    else:
                        row_dict[key] = val1
                else:
                    row_dict[key] = ""
            elif key in col_map:
                col_idx = col_map[key]
                cell_value = row.get(col_idx, "")
                row_dict[key] = str(cell_value).strip() if pd.notna(cell_value) else ""
            else:
                row_dict[key] = ""
        
        # 데이터 정리: 등기명의인에 다른 정보가 섞여있는 경우 분리
        if "등기명의인" in row_dict:
            owner_text = str(row_dict["등기명의인"]).strip()
            
            # 주민등록번호 분리
            if "(주민)등록번호" in col_keywords:
                jumin = extract_jumin_number(owner_text)
                if jumin:
                    row_dict["(주민)등록번호"] = jumin
                    owner_text = owner_text.replace(jumin, "").strip()
            
            # 지분 정보 분리
            if "최종지분" in col_keywords and not row_dict.get("최종지분"):
                extracted_jibun = extract_jibun(owner_text)
                if extracted_jibun:
                    row_dict["최종지분"] = extracted_jibun
                    owner_text = owner_text.replace(extracted_jibun, "").strip()
            
            # 주소 정보 분리
            if "주소" in col_keywords and not row_dict.get("주소"):
                if is_address_pattern(owner_text):
                    # 이름과 주소를 분리하려고 시도
                    parts = owner_text.split()
                    if len(parts) > 1:
                        # 첫 번째 부분이 이름이고 나머지가 주소일 가능성
                        possible_name = parts[0]
                        possible_address = " ".join(parts[1:])
                        if is_address_pattern(possible_address):
                            row_dict["등기명의인"] = possible_name.replace(" ", "")  # 이름 띄어쓰기 제거
                            row_dict["주소"] = possible_address
                            continue
            
            # 정리된 등기명의인 설정 (띄어쓰기 제거)
            row_dict["등기명의인"] = owner_text.replace(" ", "")
            
        rows.append(row_dict)
    
    return pd.DataFrame(rows)

def find_keyword_header(section, col_keywords, max_search_rows=15):
    section = section.fillna("").astype(str)
    for i in range(min(max_search_rows, len(section))):
        row = section.iloc[i]
        match_count = sum(any(keyword_match_exact(cell, kw) for cell in row) for kw in col_keywords)
        if match_count >= 3:
            return i, row
    return None, None

def find_col_index(header_row, keyword):
    for idx, val in header_row.items():
        if keyword_match_exact(val, keyword):
            return idx
    return None

# 소유권사항 (갑구)와 에서 필요한 열 추출
def extract_precise_named_cols(section, col_keywords):
    # 셀 병합을 하지 않고 원본 섹션 사용
    section = section.copy()
    # always use first row as header
    header_row = merge_split_headers(section.iloc[0])
    start_row = 1
    
    col_map = {}
    for key in col_keywords:
        idx = find_col_index(header_row, key)
        # fallback to partial match if exact failed
        if idx is None:
            for i, val in header_row.items():
                if keyword_match_partial(val, key):
                    idx = i
                    break
        if idx is not None:
            col_map[key] = idx

    if not col_map:
       # 모든 컬럼에 대해 빈 값을 생성하고, 첫번째 컬럼에만 "기록없음" 표시
       result = pd.DataFrame(columns=col_keywords)
       result.loc[0] = [""] * len(col_keywords)
       result.iloc[0, 0] = "기록없음"
       return result

    rows = []
    for i in range(start_row, len(section)):
        row = section.iloc[i]
        row_dict = {}
        for key in col_keywords:
            if key in col_map:
                # 해당 열의 정확한 인덱스에서만 값 가져오기
                col_idx = col_map[key]
                if col_idx < len(row):
                    cell_value = row.iloc[col_idx]
                    row_dict[key] = str(cell_value).strip() if pd.notna(cell_value) else ""
                else:
                    row_dict[key] = ""
            else:
                row_dict[key] = ""
        rows.append(row_dict)
    return pd.DataFrame(rows)
def merge_same_row_if_amount_separated(df):
    df = df.copy()
    for i in range(len(df) - 1):
        row = df.iloc[i]
        main = str(row["주요등기사항"])

        if "채권최고액" in main:
            # 현재 행과 다음 행 모두 병합 텍스트 구성
            combined_row = list(row.values) + list(df.iloc[i + 1].values)
            combined_text = " ".join(str(x) for x in combined_row if pd.notnull(x))

            # 금액 패턴 추출
            match = re.search(r"금[\d,]+원", combined_text)
            if match and match.group(0) not in main:
                df.at[i, "주요등기사항"] = main + " " + match.group(0)
    return df
def is_jumin_number(text):
    """
    주민등록번호 패턴을 확인하는 함수
    예: 123456-1234567 또는 123456-*******
    """
    if not isinstance(text, str):
        return False
    
    # 주민등록번호 패턴 (숫자6자리-숫자또는*)
    pattern = re.compile(r'\d{6}-[\d\*]+')
    return bool(re.search(pattern, text))

def extract_jumin_number(text):
    """
    문자열에서 주민등록번호 패턴을 추출
    """
    if not isinstance(text, str):
        return ""
    
    pattern = re.compile(r'\d{6}-[\d\*]+')
    match = re.search(pattern, text)
    return match.group(0) if match else ""

def is_jibun_pattern(text):
    """
    최종지분 패턴을 확인하는 함수
    예: 1/2, 50%, 3분의 1, 공유1/3, 단독소유 등
    """
    if not isinstance(text, str):
        return False
    
    # 텍스트가 비어있으면 지분 패턴 아님
    if not text.strip():
        return False
    
    # "단독소유" 키워드 확인
    if "단독소유" in text or "단독" in text:
        return True
    
    # 분수 패턴 (예: 1/2, 1/3, 공유1/3 등)
    pattern1 = re.compile(r'(?:공유)?[\d]+[/][\d]+')
    # 퍼센트 패턴 (예: 50%, 33.3% 등)
    pattern2 = re.compile(r'[\d]+[.]?[\d]*\s*%')
    # '분의' 패턴 (예: 3분의 1, 2분의 1 등)
    pattern3 = re.compile(r'[\d]+\.?[\d]*\s*분\s*의\s*[\d]+\.?[\d]*')
    # 분의 패턴 - 띄어쓰기 없는 경우 (예: 10139.94분의845.0298)
    pattern4 = re.compile(r'[\d]+\.?[\d]*분의[\d]+\.?[\d]*')
    
    return (bool(re.search(pattern1, text)) or 
            bool(re.search(pattern2, text)) or 
            bool(re.search(pattern3, text)) or 
            bool(re.search(pattern4, text)))

def is_address_pattern(text):
    """
    주소 패턴을 확인하는 함수
    """
    if not isinstance(text, str):
        return False
    
    # "단독소유" 키워드가 있으면 주소가 아님
    if "단독소유" in text or "단독" in text:
        return False
    
    # 주소에 흔히 포함되는 키워드
    address_keywords = ['시', '도', '군', '구', '읍', '면', '동', '로', '길', '아파트', '빌라', '번지']
    text_no_space = re.sub(r'\s+', '', text)
    
    for kw in address_keywords:
        if kw in text_no_space:
            return True
            
    return False

def extract_jibun(text):
    """
    문자열에서 지분 패턴 추출
    """
    if not isinstance(text, str):
        return ""
    
    # "단독소유" 키워드 확인
    if "단독소유" in text:
        return "단독소유"
    elif "단독" in text and len(text.strip()) < 10:  # "단독" 단어만 있고 길이가 짧은 경우
        return "단독소유"
    
    # 분수 패턴 (예: 1/2, 1/3, 공유1/3 등)
    pattern1 = re.compile(r'(?:공유)?[\d]+[/][\d]+')
    # 퍼센트 패턴 (예: 50%, 33.3% 등)
    pattern2 = re.compile(r'[\d]+[.]?[\d]*\s*%')
    # '분의' 패턴 - 띄어쓰기 있는 경우 (예: 3분의 1, 10139.94분 의 845.0298)
    pattern3 = re.compile(r'[\d]+\.?[\d]*\s*분\s*의\s*[\d]+\.?[\d]*')
    # 분의 패턴 - 띄어쓰기 없는 경우 (예: 10139.94분의845.0298)
    pattern4 = re.compile(r'[\d]+\.?[\d]*분의[\d]+\.?[\d]*')
    
    # 각 패턴 순서대로 확인
    match1 = re.search(pattern1, text)
    if match1:
        return match1.group(0)
    
    match2 = re.search(pattern2, text)
    if match2:
        return match2.group(0)
    
    match3 = re.search(pattern3, text)
    if match3:
        return match3.group(0)
    
    match4 = re.search(pattern4, text)
    if match4:
        return match4.group(0)
    
    return ""

def extract_ownership_type(owner_name):
    """
    등기명의인 문자열에서 소유구분 정보(소유자, 공유자 등)를 추출하는 함수
    """
    if not isinstance(owner_name, str):
        return "", owner_name
    
    # (소유자), (공유자) 패턴 찾기
    pattern = r'\((소유자|공유자)\)'
    match = re.search(pattern, owner_name)
    
    if match:
        ownership_type = match.group(1)  # '소유자' 또는 '공유자' 추출
        clean_name = owner_name.replace(match.group(0), "").strip()  # 패턴 제거
        return ownership_type, clean_name
    else:
        return "", owner_name

def extract_land_type(df):
    """
    엑셀 파일에서 토지 지목 정보를 추출하는 함수
    """
    land_type = ""
    # 더 구체적이고 긴 단어가 먼저 검사되도록 정렬
    land_types = ["공장용지", "잡종지", "염전", "도로", "임야", "유지", "하천", "구거", "제방", "양어장","전", "답", "대","광천지","수도용지","제방","염전","과수원","목장용지","학교용지","종교용지","주차장","주유소","창고용지","철도용지","공원","묘지","체육용지","유원지","사적지","잡종지"]
    
    # 1. 주요 등기사항 요약 섹션에서 토지 지목 추출 시도 (최우선)
    summary_row_idx = None
    for i in range(len(df)):
        row_text = " ".join(str(cell) for cell in df.iloc[i] if pd.notna(cell))
        if "주요 등기사항 요약" in row_text or "주요등기사항요약" in re.sub(r'\s+', '', row_text):
            summary_row_idx = i
            break
    
    if summary_row_idx is not None:
        # 요약 섹션 이후 토지 정보 검색
        for i in range(summary_row_idx + 1, min(summary_row_idx + 10, len(df))):
            row_text = " ".join(str(cell) for cell in df.iloc[i] if pd.notna(cell))
            if "[토지]" in row_text:
                # 지목 정보를 더 정확하게 추출
                for lt in land_types:
                    # [토지] 다음에 오는 지목 정보 찾기
                    pattern = r'\[토지\][^가-힣]*' + lt + r'(?:\s|$|[^가-힣])'
                    if re.search(pattern, row_text):
                        return lt
                    # 간단한 패턴도 확인
                    if lt in row_text and "[토지]" in row_text:
                        # 주변 문맥 확인하여 실제 지목인지 판단
                        lt_index = row_text.find(lt)
                        land_index = row_text.find("[토지]")
                        if abs(lt_index - land_index) < 50:  # 50자 이내에 있으면 관련성 있음
                            return lt
    
    # 2. 파일 식별자에서 지목 정보 추출 시도
    identifier = extract_identifier(df)
    if "[토지]" in identifier:
        # 정확한 매칭을 위한 패턴: 앞뒤로 공백이나 문장 끝인 경우만 매칭
        for lt in land_types:
            pattern = r'(^|\s|[^가-힣])' + lt + r'($|\s|[^가-힣])'
            if re.search(pattern, identifier):
                land_type = lt
                break
                
        # 정확한 매칭이 안 된 경우 부분 매칭으로 시도 (단, 더 엄격하게)
        if not land_type:
            for lt in land_types:
                if lt in identifier and "[토지]" in identifier:
                    # 지목이 [토지] 근처에 있는지 확인
                    lt_index = identifier.find(lt)
                    land_index = identifier.find("[토지]")
                    if abs(lt_index - land_index) < 30:  # 30자 이내
                        land_type = lt
                        break
    
    # 3. 데이터프레임 전체에서 찾기 (더 신중하게)
    if not land_type:
        for i in range(len(df)):
            row_text = " ".join(str(cell) for cell in df.iloc[i] if pd.notna(cell))
            
            # [토지] 키워드가 있는 행 우선 검색
            if "[토지]" in row_text:
                for lt in land_types:
                    pattern = r'(^|\s|[^가-힣])' + lt + r'($|\s|[^가-힣])'
                    if re.search(pattern, row_text):
                        return lt
                
                # 정확한 매칭이 안 되면 부분 매칭 시도 (단, [토지] 근처에서만)
                for lt in land_types:
                    if lt in row_text:
                        lt_index = row_text.find(lt)
                        land_index = row_text.find("[토지]")
                        if abs(lt_index - land_index) < 30:
                            return lt
            
            # 지목과 면적이 함께 나오는 패턴 찾기
            for lt in land_types:
                if lt in row_text and ("㎡" in row_text or "m²" in row_text):
                    # 지목과 면적이 같은 행에 있으면 실제 지목일 가능성 높음
                    return lt
    
    return land_type if land_type else ""

def extract_land_area(df):
    """
    엑셀 파일에서 토지면적 정보를 추출하는 함수
    다양한 형식의 면적 표기를 인식
    """
    area = ""
    land_types = ["염전", "도로", "임야", "유지", "답", "전", "대", "공장용지", "잡종지", "하천", "구거", "제방", "양어장"]
    
    # 주요 등기사항 요약 섹션에서 면적 추출 시도
    summary_row_idx = None
    for i in range(len(df)):
        row_text = " ".join(str(cell) for cell in df.iloc[i] if pd.notna(cell))
        if "주요 등기사항 요약" in row_text or "주요등기사항요약" in re.sub(r'\s+', '', row_text):
            summary_row_idx = i
            break
    
    if summary_row_idx is not None:
        # 요약 섹션 이후 토지 정보 검색
        for i in range(summary_row_idx + 1, min(summary_row_idx + 10, len(df))):
            row_text = " ".join(str(cell) for cell in df.iloc[i] if pd.notna(cell))
            if "[토지]" in row_text:
                area_match = re.search(r'(\d[\d,\.]*)\s*[㎡m²]', row_text)
                if area_match:
                    return area_match.group(1).replace(',', '')
    
    # 이하 기존 추출 방법 (위 방법이 실패한 경우 실행)
    # 파일 식별자에서 면적 추출 시도
    identifier = extract_identifier(df)
    if "[토지]" in identifier:
        # 면적 패턴 찾기: "[토지]" 문장 내에서 숫자 + ㎡ 또는 m² 패턴
        area_match = re.search(r'(\d[\d,\.]*)\s*[㎡m²]', identifier)
        if area_match:
            return area_match.group(1).replace(',', '')
    
    # 데이터프레임 전체에서 찾기
    for i in range(len(df)):
        row_text = " ".join(str(cell) for cell in df.iloc[i] if pd.notna(cell))
        
        # 토지종류가 있는 행에서 면적 패턴 찾기
        if any(land_type in row_text for land_type in land_types):
            # 면적 패턴: 숫자 + ㎡ 또는 m² 패턴
            area_match = re.search(r'(\d[\d,\.]*)\s*[㎡m²]', row_text)
            if area_match:
                area = area_match.group(1).replace(',', '')
                break
            
        # "[토지]" 패턴이 있는 행에서 찾기
        if "[토지]" in row_text:
            area_match = re.search(r'(\d[\d,\.]*)\s*[㎡m²]', row_text)
            if area_match:
                area = area_match.group(1).replace(',', '')
                break
    
    return area

def check_san_in_address(address):
    """
    토지주소에 '산'이 있는지 확인하는 함수
    '산'이 숫자 앞에 있으면 'O', 아니면 'X'
    """
    if not isinstance(address, str):
        return ''
    
    # 주소에서 마지막 부분을 가져오기
    parts = address.split()
    if not parts:
        return ''
    
    # 주소의 마지막 부분에서 '산' 다음에 숫자가 오는 패턴 확인
    import re
    for part in parts:
        if re.search(r'산\d+', part) or re.search(r'산\s*\d+', part):
            return '산'
    return ''

def extract_right_holders(df):
    """
    주요등기사항에서 근저당권자와 지상권자 정보를 추출하고, 
    원본 텍스트에서 해당 정보를 제거하는 함수
    """
    df = df.copy()
    df["근저당권자"] = ""
    df["지상권자"] = ""
    
    for idx, row in df.iterrows():
        if "주요등기사항" not in row or pd.isna(row["주요등기사항"]):
            continue
            
        main_text = str(row["주요등기사항"])
        modified_text = main_text
        
        # 근저당권자 추출 및 제거
        mortgage_pattern = r'근저당권자\s*[:：]?\s*([^,\n]*)'
        mortgage_match = re.search(mortgage_pattern, main_text)
        if mortgage_match:
            df.at[idx, "근저당권자"] = mortgage_match.group(1).strip()
            # 전체 매치 부분을 찾아 제거 (근저당권자: XXX 형태 전체)
            full_match = mortgage_match.group(0)
            modified_text = modified_text.replace(full_match, "")
        
        # 지상권자 추출 및 제거
        surface_pattern = r'지상권자\s*[:：]?\s*([^,\n]*)'
        surface_match = re.search(surface_pattern, modified_text)
        if surface_match:
            df.at[idx, "지상권자"] = surface_match.group(1).strip()
            # 전체 매치 부분을 찾아 제거 (지상권자: XXX 형태 전체)
            full_match = surface_match.group(0)
            modified_text = modified_text.replace(full_match, "")
        
        # 수정된 텍스트 정리 (앞뒤 공백, 쉼표 정리)
        modified_text = modified_text.strip()
        modified_text = re.sub(r',\s*,', ',', modified_text)  # 연속된 쉼표 제거
        modified_text = re.sub(r'^\s*,\s*|\s*,\s*$', '', modified_text)  # 시작/끝의 쉼표 제거
        
        # 정리된 텍스트로 업데이트
        df.at[idx, "주요등기사항"] = modified_text
    
    return df

def style_header_row(ws):
    """워크시트 헤더 행을 스타일링하는 함수"""
    # 연한 초록색 배경 설정 (RGB: 230, 244, 234)
    light_green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    
    # 테두리 스타일 정의
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 첫 번째 행 (헤더) 스타일 적용
    for cell in ws[1]:
        # 중앙 정렬
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 연한 초록색 배경
        cell.fill = light_green_fill
        # 테두리 추가
        cell.border = thin_border
    
    # 헤더 행 높이 조정
    ws.row_dimensions[1].height = 25
    
    # 열 너비 자동 조정 (내용에 따라)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        # 각 셀의 내용 길이 확인
        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                max_length = max(max_length, cell_length)
            except:
                pass
        # 최소 10, 최대 50 사이로 너비 조정
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted_width

def create_grouped_headers(ws, df, group_structure):
    """
    워크시트에 그룹화된 헤더를 생성하는 함수
    group_structure: {그룹명: [컬럼명 리스트]} 형태의 딕셔너리
    """
    # 첫 번째 행 - 그룹 헤더
    row_index = 1
    col_index = 1
    
    # 연한 초록색 배경 설정 (RGB: 230, 244, 234)
    light_green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    
    # 테두리 스타일 정의
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 그룹 헤더 행 추가
    for group_name, columns in group_structure.items():
        # 그룹 이름 셀
        group_cell = ws.cell(row=row_index, column=col_index)
        group_cell.value = group_name
        group_cell.alignment = Alignment(horizontal='center', vertical='center')
        group_cell.fill = light_green_fill
        group_cell.border = thin_border
        
        # 여러 열에 걸쳐 병합
        if len(columns) > 1:
            ws.merge_cells(start_row=row_index, start_column=col_index, 
                          end_row=row_index, end_column=col_index + len(columns) - 1)
            
            # 병합된 셀에 테두리 추가 (병합 후에 모든 셀에 테두리 적용)
            for c in range(col_index, col_index + len(columns)):
                cell = ws.cell(row=row_index, column=c)
                cell.border = thin_border
        
        col_index += len(columns)
    
    # 두 번째 행 - 세부 헤더
    row_index = 2
    col_index = 1
    
    for _, columns in group_structure.items():
        for col_name in columns:
            col_cell = ws.cell(row=row_index, column=col_index)
            col_cell.value = col_name
            col_cell.alignment = Alignment(horizontal='center', vertical='center')
            col_cell.fill = light_green_fill
            col_cell.border = thin_border  # 각 열 헤더에 테두리 추가
            col_index += 1
    
    # 데이터 추가 (3번째 행부터)
    row_index = 3
    for _, row in df.iterrows():
        col_index = 1
        for _, columns in group_structure.items():
            for col_name in columns:
                cell = ws.cell(row=row_index, column=col_index)
                cell.value = row.get(col_name, "")
                # 데이터 셀에도 가벼운 테두리 추가 (선택적)
                cell.border = Border(
                    left=Side(style='thin', color='D3D3D3'),
                    right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'),
                    bottom=Side(style='thin', color='D3D3D3')
                )
                col_index += 1
        row_index += 1
    
    # 열 너비 자동 조정 (내용에 따라)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        # 각 셀의 내용 길이 확인
        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                max_length = max(max_length, cell_length)
            except:
                pass
        # 최소 10, 최대 50 사이로 너비 조정
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted_width

if run_button and uploaded_zip:
    temp_dir = tempfile.mkdtemp()
    szj_list, syg_list, djg_list = [], [], []

    with zipfile.ZipFile(uploaded_zip, "r") as z:
        z.extractall(temp_dir)

    # ✅ 하위 폴더 포함 모든 .xlsx 탐색
    excel_files = []
    for root, _, files in os.walk(temp_dir):
        for f in files:
            if f.lower().endswith(".xlsx"):
                excel_files.append(os.path.join(root, f))

    for path in excel_files:
        try:
            xls = pd.ExcelFile(path)
            df = xls.parse(xls.sheet_names[0]).fillna("")
            name = extract_identifier(df)
            
            # 토지면적과 지목 정보 추출
            land_area = extract_land_area(df)
            land_type = extract_land_type(df)

            szj_sec, has_szj = extract_section_range(df, "소유지분현황", ["소유권", "저당권"], match_fn=keyword_match_partial)
            syg_sec, has_syg = extract_section_range(df, "소유권.*사항", ["저당권"], match_fn=keyword_match_exact)
            djg_sec, has_djg = extract_section_range(df, "3.(근)저당권및전세권등(을구)", ["참고", "비고", "총계", "전산자료"], match_fn=keyword_match_exact)

            if has_szj:
                szj_df = extract_named_cols(szj_sec, ["등기명의인", "(주민)등록번호", "최종지분", "주소", "순위번호"])
                
                # 소유구분 열 추가
                szj_df["소유구분"] = ""
                
                # 데이터 후처리 - 등기명의인과 주민등록번호 정리
                for idx, row in szj_df.iterrows():
                    # 소유구분 추출
                    if pd.notna(row["등기명의인"]):
                        ownership_type, clean_name = extract_ownership_type(str(row["등기명의인"]))
                        szj_df.at[idx, "소유구분"] = ownership_type
                        szj_df.at[idx, "등기명의인"] = clean_name.replace(" ", "")  # 등기명의인 띄어쓰기 제거
                    
                    # 등기명의인에서 주민번호 패턴이 있으면 분리
                    if pd.notna(row["등기명의인"]):
                        jumin = extract_jumin_number(str(row["등기명의인"]))
                        if jumin:
                            szj_df.at[idx, "(주민)등록번호"] = jumin
                            szj_df.at[idx, "등기명의인"] = str(row["등기명의인"]).replace(jumin, "").strip().replace(" ", "")  # 띄어쓰기 제거

                    # 최종지분과 주소 추가 정리
                    address_text = str(row["주소"]).strip()
                    jibun_text = str(row["최종지분"]).strip()
                    
                    # 주소에서 단독소유 또는 지분 패턴 찾기
                    if pd.notna(row["주소"]) and is_jibun_pattern(address_text):
                        jibun_in_address = extract_jibun(address_text)
                        if jibun_in_address:
                            # 최종지분이 비어있거나, 주소에서 발견한 지분이 더 정확해 보이는 경우
                            if not jibun_text or len(jibun_in_address) > len(jibun_text):
                                szj_df.at[idx, "최종지분"] = jibun_in_address
                            # 주소에서는 지분 정보 제거
                            szj_df.at[idx, "주소"] = address_text.replace(jibun_in_address, "").strip()
                    
                    # 최종지분에 주소 패턴 찾기
                    if pd.notna(row["최종지분"]) and is_address_pattern(jibun_text):
                        # 주소 필드가 비어있거나 최종지분의 텍스트가 더 길면(상세 주소일 가능성)
                        if not address_text or (len(jibun_text) > len(address_text)):
                            szj_df.at[idx, "주소"] = jibun_text
                            szj_df.at[idx, "최종지분"] = ""
                
                # 마지막 검증 - 단독소유 확인
                for idx, row in szj_df.iterrows():
                    address_text = str(row["주소"]).strip()
                    if "단독" in address_text and "단독소유" not in str(row["최종지분"]):
                        # 단독 텍스트가 주소에 있고 최종지분에 없으면 이동
                        szj_df.at[idx, "최종지분"] = "단독소유"
                        # 주소에서는 '단독' 또는 '단독소유' 제거
                        szj_df.at[idx, "주소"] = re.sub(r'단독(?:소유)?', '', address_text).strip()
                
                # 최종지분에서 주소 정보 제거하기
                for idx, row in szj_df.iterrows():
                    jibun_text = str(row["최종지분"]).strip()
                    
                    # 최종지분에서 지분 패턴 추출
                    if jibun_text and pd.notna(row["최종지분"]):
                        if "단독소유" in jibun_text or "단독" in jibun_text and len(jibun_text) < 10:
                            # 단독소유는 그대로 유지
                            szj_df.at[idx, "최종지분"] = "단독소유"
                        else:
                            # 지분 패턴만 추출
                            extracted_jibun = extract_jibun(jibun_text)
                            if extracted_jibun:
                                szj_df.at[idx, "최종지분"] = extracted_jibun
                            else:
                                # 주소 패턴 확인 후 주소라면 해당 필드를 비움
                                if is_address_pattern(jibun_text):
                                    if str(row["주소"]).strip() == "":
                                        szj_df.at[idx, "주소"] = jibun_text
                                    szj_df.at[idx, "최종지분"] = ""
                
                # 토지면적 열 추가
                szj_df["지목"] = land_type      # 지목 열 추가
                szj_df["토지면적"] = land_area
                
                # 소유면적 계산 및 열 추가
                szj_df["지분면적"] = None
                for idx, row in szj_df.iterrows():
                    try:
                        jibun_decimal = convert_jibun_to_decimal(row["최종지분"])
                        if jibun_decimal is not None and pd.notna(row["토지면적"]) and row["토지면적"]:
                            land_area_value = float(str(row["토지면적"]).replace(',', ''))
                            ownership_area = land_area_value * jibun_decimal
                            szj_df.at[idx, "지분면적"] = f"{ownership_area:.4f}"
                    except Exception as e:
                        pass  # 변환 중 오류 발생시 None 값 유지
                
                # 최종지분 수치화 열 추가
                szj_df["최종지분 수치화"] = None
                for idx, row in szj_df.iterrows():
                    try:
                        jibun_decimal = convert_jibun_to_decimal(row["최종지분"])
                        if jibun_decimal is not None:
                            szj_df.at[idx, "최종지분 수치화"] = jibun_decimal
                    except Exception as e:
                        pass  # 변환 중 오류 발생시 None 값 유지
                
                # 열 순서 재배치
                szj_df.insert(0, "토지주소", name)
                columns = ["토지주소", "등기명의인", "소유구분", "(주민)등록번호", "주소", "순위번호", "최종지분", "최종지분 수치화", "지목", "토지면적", "지분면적"]
                szj_df = szj_df[columns]
                szj_df["그룹정보"] = "있음"  # 그룹 헤더를 사용할 데이터 플래그
                szj_list.append(szj_df)
            else:
                # "기록없음" 케이스에도 동일한 컬럼 구조 유지
                szj_list.append(pd.DataFrame([[name, "기록없음", "", "", "", "", "", "", land_type, land_area, "", "없음"]], 
                                             columns=["토지주소", "등기명의인", "소유구분", "(주민)등록번호", "주소", "순위번호", "최종지분", "최종지분 수치화", "지목", "토지면적", "지분면적", "그룹정보"]))

            if has_syg:
                syg_df = extract_precise_named_cols(syg_sec, ["순위번호", "등기목적", "접수정보", "주요등기사항", "대상소유자"])
                syg_df.insert(0, "토지주소", name)
                syg_list.append(syg_df)
            else:
                syg_list.append(pd.DataFrame([[name, "기록없음"]], columns=["토지주소", "순위번호"]))

            if has_djg:
                djg_df = extract_precise_named_cols(djg_sec, ["순위번호", "등기목적", "접수정보", "주요등기사항", "대상소유자"])
                
                # 빈 행 제거 - 빈 문자열을 NA로 변환 후 모든 값이 NA인 행 제거
                djg_df = djg_df.replace('', pd.NA)
                djg_df = djg_df.dropna(how='all')
                
                # 공백만 있는 행도 제거 (문자열을 trim한 후 빈 문자열인지 확인)
                mask = ~djg_df.astype(str).apply(lambda row: row.str.strip().eq('').all(), axis=1)
                djg_df = djg_df[mask].reset_index(drop=True)
                
                # 빈 값을 다시 빈 문자열로 변환
                djg_df = djg_df.fillna('')
                
                # "대상소유자" 컬럼에서 모든 띄어쓰기 제거
                if "대상소유자" in djg_df.columns:
                    djg_df["대상소유자"] = djg_df["대상소유자"].astype(str).str.replace(" ", "")
                
                djg_df = merge_same_row_if_amount_separated(djg_df)
                djg_df = trim_after_reference_note(djg_df)
                djg_df = extract_right_holders(djg_df)
                djg_df.insert(0, "토지주소", name)
                
                djg_list.append(djg_df)
            else:
                # 빈 데이터프레임에도 모든 열 포함 - 기록유무 열 제거
                djg_list.append(pd.DataFrame([[name, "기록없음", "", "", "", "", "", ""]], 
                                           columns=["토지주소", "순위번호", "등기목적", "접수정보", "주요등기사항", "대상소유자", "근저당권자", "지상권자"]))

        except Exception as e:
            pass  # 또는 logging.warning(...) 등으로 로깅만
    wb = Workbook()
    for sheetname, data in zip(
        ["1. 소유지분현황 (갑구)", "2. 소유권사항 (갑구)", "3. 저당권사항 (을구)"],
        [szj_list, syg_list, djg_list]
    ):
        ws = wb.create_sheet(title=sheetname)
        if data and sheetname == "1. 소유지분현황 (갑구)":
            df = pd.concat(data, ignore_index=True)
            
            # "산" 열 추가
            df["산"] = df["토지주소"].apply(check_san_in_address)
            
            # 열 순서 재배치 - "토지주소" 다음에 "산" 위치
            cols = df.columns.tolist()
            cols.remove("산")
            idx = cols.index("토지주소")
            cols.insert(idx + 1, "산")
            df = df[cols]
            
            # 소유지분현황(갑구) 시트에는 그룹 헤더 적용
            if any(df["그룹정보"] == "있음"):
                # 그룹 구조 정의 - "산" 열 추가
                group_structure = {
                    "토지주소": ["토지주소", "산"],
                    "소유자": ["등기명의인", "소유구분", "(주민)등록번호", "주소", "순위번호"],
                    "토지": ["최종지분", "최종지분 수치화", "지목", "토지면적", "지분면적"]
                }
                df = df.drop(columns=["그룹정보"])  # 그룹정보 열 제거
                create_grouped_headers(ws, df, group_structure)
            else:
                df = df.drop(columns=["그룹정보"])  # 그룹정보 열 제거
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                # 헤더 행 스타일 적용
                style_header_row(ws)
        elif data:
            df = pd.concat(data, ignore_index=True)
            df.reset_index(drop=True, inplace=True)
            
            if sheetname == "3. 저당권사항 (을구)":
                if "순위번호" in df.columns and "등기목적" in df.columns:
                    df = df.rename(columns={"순위번호": "기록유무"})
                    # 기록유무에 등기목적 값만 표시 (등기목적이 비어있으면 "기록없음")
                    df["기록유무"] = df["등기목적"].apply(
                        lambda x: x if pd.notna(x) and str(x).strip() and str(x).strip() != "기록없음"
                        else "기록없음"
                    )
                    df = df.drop(columns=["등기목적"])
            
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            # Headers styling
            style_header_row(ws)
        else:
            ws.append(["기록없음"])
            # 데이터가 없는 경우에도 헤더 스타일 적용
            style_header_row(ws)

    wb.remove(wb["Sheet"])
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        st.success("✅ 분석 완료! 다운로드 버튼을 클릭하세요.")
        st.download_button("📥 결과 다운로드", data=open(tmp.name, "rb"), file_name="등기사항_통합_시트별구성.xlsx")
